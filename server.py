"""
Sakhi Sang – Sunday Class Attendance & Devotee Management System
Python/Flask backend using SQLite
"""
import os
import sqlite3
import json
import io
from datetime import datetime, date, timedelta
from flask import Flask, request, jsonify, send_file, send_from_directory
from werkzeug.utils import secure_filename
import openpyxl

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
DB_PATH    = os.path.join(BASE_DIR, 'devotees.db')
PUBLIC_DIR = os.path.join(BASE_DIR, 'public')
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')

os.makedirs(UPLOAD_DIR, exist_ok=True)

app = Flask(__name__, static_folder=PUBLIC_DIR, static_url_path='')


# ── DATABASE ──────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    with get_db() as db:
        db.executescript("""
        CREATE TABLE IF NOT EXISTS devotees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            mobile TEXT,
            address TEXT,
            dob TEXT,
            date_of_joining TEXT,
            chanting_rounds INTEGER DEFAULT 0,
            kanthi INTEGER DEFAULT 0,
            gopi_dress INTEGER DEFAULT 0,
            team_name TEXT,
            devotee_status TEXT DEFAULT 'Expected to be Serious',
            facilitator TEXT,
            reference_by TEXT,
            calling_by TEXT,
            lifetime_attendance INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            inactivity_flag INTEGER DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS attendance_sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_date TEXT NOT NULL UNIQUE,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS attendance_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER NOT NULL,
            devotee_id INTEGER NOT NULL,
            is_new_devotee INTEGER DEFAULT 0,
            marked_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (session_id) REFERENCES attendance_sessions(id),
            FOREIGN KEY (devotee_id) REFERENCES devotees(id),
            UNIQUE(session_id, devotee_id)
        );
        CREATE TABLE IF NOT EXISTS calling_status (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            devotee_id INTEGER NOT NULL,
            week_date TEXT NOT NULL,
            coming_status TEXT DEFAULT 'Maybe',
            calling_notes TEXT,
            called_by TEXT,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (devotee_id) REFERENCES devotees(id),
            UNIQUE(devotee_id, week_date)
        );
        CREATE TABLE IF NOT EXISTS profile_changes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            devotee_id INTEGER NOT NULL,
            field_name TEXT NOT NULL,
            old_value TEXT,
            new_value TEXT,
            changed_at TEXT DEFAULT CURRENT_TIMESTAMP,
            changed_by TEXT DEFAULT 'Coordinator',
            FOREIGN KEY (devotee_id) REFERENCES devotees(id)
        );
        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_name TEXT NOT NULL,
            event_date TEXT,
            description TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS event_devotees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_id INTEGER NOT NULL,
            devotee_id INTEGER NOT NULL,
            added_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (event_id) REFERENCES events(id),
            FOREIGN KEY (devotee_id) REFERENCES devotees(id),
            UNIQUE(event_id, devotee_id)
        );
        """)


init_db()


# ── HELPERS ───────────────────────────────────────────────────────────────────
def today_str():
    return date.today().isoformat()


def current_sunday():
    d = date.today()
    return (d - timedelta(days=d.weekday() + 1) if d.weekday() != 6 else d).isoformat()


def row_to_dict(row):
    return dict(row) if row else None


def rows_to_list(rows):
    return [dict(r) for r in rows]


def update_inactivity_flags():
    with get_db() as db:
        sessions = db.execute("SELECT id FROM attendance_sessions ORDER BY session_date DESC LIMIT 3").fetchall()
        if len(sessions) < 3:
            return
        sid1, sid2, sid3 = sessions[0]['id'], sessions[1]['id'], sessions[2]['id']
        devotees = db.execute("SELECT id FROM devotees WHERE is_active=1").fetchall()
        for d in devotees:
            attended = db.execute(
                "SELECT COUNT(*) as c FROM attendance_records WHERE devotee_id=? AND session_id IN (?,?,?)",
                (d['id'], sid1, sid2, sid3)
            ).fetchone()['c']
            db.execute("UPDATE devotees SET inactivity_flag=? WHERE id=?", (1 if attended == 0 else 0, d['id']))


# ── STATIC ────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return send_from_directory(PUBLIC_DIR, 'index.html')


# ══════════════════════════════════════════════════════════════════════════════
# DEVOTEES
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/devotees', methods=['GET'])
def get_devotees():
    search     = request.args.get('search', '')
    team       = request.args.get('team', '')
    calling_by = request.args.get('calling_by', '')
    status     = request.args.get('status', '')
    q = "SELECT * FROM devotees WHERE is_active=1"
    params = []
    if search:
        q += " AND (name LIKE ? OR mobile LIKE ?)"; params += [f'%{search}%', f'%{search}%']
    if team:
        q += " AND team_name=?"; params.append(team)
    if calling_by:
        q += " AND calling_by=?"; params.append(calling_by)
    if status:
        q += " AND devotee_status=?"; params.append(status)
    q += " ORDER BY name ASC"
    with get_db() as db:
        return jsonify(rows_to_list(db.execute(q, params).fetchall()))


@app.route('/api/devotees/meta/calling-persons', methods=['GET'])
def calling_persons():
    with get_db() as db:
        rows = db.execute("SELECT DISTINCT calling_by FROM devotees WHERE calling_by IS NOT NULL AND calling_by!='' ORDER BY calling_by").fetchall()
        return jsonify([r['calling_by'] for r in rows])


@app.route('/api/devotees/import', methods=['POST'])
def import_devotees():
    if 'file' not in request.files:
        return jsonify({'error': 'No file'}), 400
    f = request.files['file']
    path = os.path.join(UPLOAD_DIR, secure_filename(f.filename or 'import.xlsx'))
    f.save(path)
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
        imported = skipped = 0
        errors = []
        with get_db() as db:
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    d = {headers[j]: (str(v).strip() if v is not None else '') for j, v in enumerate(row)}
                    name   = d.get('Name') or d.get('name') or ''
                    mobile = d.get('Mobile') or d.get('mobile') or d.get('Phone') or ''
                    if not name:
                        skipped += 1; continue
                    if mobile:
                        ex = db.execute("SELECT id FROM devotees WHERE mobile=?", (mobile,)).fetchone()
                        if ex:
                            skipped += 1; continue
                    cur = db.execute(
                        "INSERT OR IGNORE INTO devotees (name,mobile,address,dob,date_of_joining,chanting_rounds,kanthi,gopi_dress,team_name,devotee_status,facilitator,reference_by,calling_by) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        (name, mobile or None,
                         d.get('Address') or None, d.get('DOB') or d.get('Date of Birth') or None,
                         d.get('Date of Joining') or None,
                         int(d.get('Chanting Rounds') or 0),
                         1 if d.get('Kanthi') == 'Yes' else 0,
                         1 if d.get('Gopi Dress') == 'Yes' else 0,
                         d.get('Team') or d.get('team_name') or None,
                         d.get('Status') or 'Expected to be Serious',
                         d.get('Facilitator') or None, d.get('Reference') or None,
                         d.get('Calling By') or None)
                    )
                    if cur.rowcount:
                        imported += 1
                    else:
                        skipped += 1
                except Exception as e:
                    errors.append(f"Row {i}: {str(e)}")
        return jsonify({'imported': imported, 'skipped': skipped, 'errors': errors})
    finally:
        try: os.remove(path)
        except: pass


@app.route('/api/devotees/<int:did>', methods=['GET'])
def get_devotee(did):
    with get_db() as db:
        d = db.execute("SELECT * FROM devotees WHERE id=?", (did,)).fetchone()
        if not d: return jsonify({'error': 'Not found'}), 404
        return jsonify(row_to_dict(d))


@app.route('/api/devotees', methods=['POST'])
def create_devotee():
    data = request.json or {}
    name   = (data.get('name') or '').strip()
    mobile = (data.get('mobile') or '').strip() or None
    with get_db() as db:
        if mobile:
            ex = db.execute("SELECT id,name FROM devotees WHERE mobile=?", (mobile,)).fetchone()
            if ex:
                return jsonify({'error':'Duplicate','message':f"Mobile already registered to {ex['name']}","existingId":ex['id']}), 409
        if name:
            exn = db.execute("SELECT id,name FROM devotees WHERE LOWER(TRIM(name))=LOWER(TRIM(?))", (name,)).fetchone()
            if exn:
                return jsonify({'error':'DuplicateName','message':f"Name already exists: {exn['name']}","existingId":exn['id'],"confirm":True}), 409
        cur = db.execute(
            "INSERT INTO devotees (name,mobile,address,dob,date_of_joining,chanting_rounds,kanthi,gopi_dress,team_name,devotee_status,facilitator,reference_by,calling_by) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (name, mobile, data.get('address') or None, data.get('dob') or None,
             data.get('date_of_joining') or None, int(data.get('chanting_rounds') or 0),
             int(data.get('kanthi') or 0), int(data.get('gopi_dress') or 0),
             data.get('team_name') or None, data.get('devotee_status') or 'Expected to be Serious',
             data.get('facilitator') or None, data.get('reference_by') or None,
             data.get('calling_by') or None)
        )
        new = db.execute("SELECT * FROM devotees WHERE id=?", (cur.lastrowid,)).fetchone()
        return jsonify(row_to_dict(new)), 201


@app.route('/api/devotees/force', methods=['POST'])
def force_create_devotee():
    data = request.json or {}
    name   = (data.get('name') or '').strip()
    mobile = (data.get('mobile') or '').strip() or None
    with get_db() as db:
        cur = db.execute(
            "INSERT INTO devotees (name,mobile,address,dob,date_of_joining,chanting_rounds,kanthi,gopi_dress,team_name,devotee_status,facilitator,reference_by,calling_by) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (name, mobile, data.get('address') or None, data.get('dob') or None,
             data.get('date_of_joining') or None, int(data.get('chanting_rounds') or 0),
             int(data.get('kanthi') or 0), int(data.get('gopi_dress') or 0),
             data.get('team_name') or None, data.get('devotee_status') or 'Expected to be Serious',
             data.get('facilitator') or None, data.get('reference_by') or None,
             data.get('calling_by') or None)
        )
        new = db.execute("SELECT * FROM devotees WHERE id=?", (cur.lastrowid,)).fetchone()
        return jsonify(row_to_dict(new)), 201


@app.route('/api/devotees/<int:did>', methods=['PUT'])
def update_devotee(did):
    with get_db() as db:
        existing = db.execute("SELECT * FROM devotees WHERE id=?", (did,)).fetchone()
        if not existing: return jsonify({'error':'Not found'}), 404
        ex = dict(existing)
        u = request.json or {}
        tracked = ['name','mobile','chanting_rounds','kanthi','gopi_dress','team_name','devotee_status','facilitator','reference_by','calling_by']
        for f in tracked:
            if f in u and str(u[f]) != str(ex.get(f) or ''):
                db.execute("INSERT INTO profile_changes (devotee_id,field_name,old_value,new_value,changed_by) VALUES (?,?,?,?,?)",
                           (did, f, ex.get(f), u[f], u.get('_changed_by','Coordinator')))
        db.execute("""UPDATE devotees SET name=?,mobile=?,address=?,dob=?,date_of_joining=?,chanting_rounds=?,
                      kanthi=?,gopi_dress=?,team_name=?,devotee_status=?,facilitator=?,reference_by=?,calling_by=?,
                      updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                   (u.get('name',ex['name']), u.get('mobile',ex['mobile']), u.get('address',ex['address']),
                    u.get('dob',ex['dob']), u.get('date_of_joining',ex['date_of_joining']),
                    u.get('chanting_rounds',ex['chanting_rounds']),
                    u.get('kanthi',ex['kanthi']), u.get('gopi_dress',ex['gopi_dress']),
                    u.get('team_name',ex['team_name']), u.get('devotee_status',ex['devotee_status']),
                    u.get('facilitator',ex['facilitator']), u.get('reference_by',ex['reference_by']),
                    u.get('calling_by',ex['calling_by']), did))
        updated = db.execute("SELECT * FROM devotees WHERE id=?", (did,)).fetchone()
        return jsonify(row_to_dict(updated))


@app.route('/api/devotees/<int:did>', methods=['DELETE'])
def delete_devotee(did):
    with get_db() as db:
        db.execute("UPDATE devotees SET is_active=0 WHERE id=?", (did,))
        return jsonify({'success': True})


@app.route('/api/devotees/<int:did>/history', methods=['GET'])
def devotee_history(did):
    with get_db() as db:
        rows = db.execute("SELECT * FROM profile_changes WHERE devotee_id=? ORDER BY changed_at DESC", (did,)).fetchall()
        return jsonify(rows_to_list(rows))


# ══════════════════════════════════════════════════════════════════════════════
# SESSIONS / ATTENDANCE
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/sessions/today', methods=['GET'])
def get_today_session():
    t = today_str()
    with get_db() as db:
        s = db.execute("SELECT * FROM attendance_sessions WHERE session_date=?", (t,)).fetchone()
        if not s:
            cur = db.execute("INSERT INTO attendance_sessions (session_date) VALUES (?)", (t,))
            s = db.execute("SELECT * FROM attendance_sessions WHERE id=?", (cur.lastrowid,)).fetchone()
        return jsonify(row_to_dict(s))


@app.route('/api/sessions', methods=['GET'])
def get_sessions():
    with get_db() as db:
        rows = db.execute("SELECT * FROM attendance_sessions ORDER BY session_date DESC LIMIT 30").fetchall()
        return jsonify(rows_to_list(rows))


@app.route('/api/sessions/<int:sid>/stats', methods=['GET'])
def session_stats(sid):
    week = current_sunday()
    with get_db() as db:
        target   = db.execute("SELECT COUNT(*) as c FROM calling_status WHERE week_date=? AND coming_status='Yes'", (week,)).fetchone()['c']
        present  = db.execute("SELECT COUNT(*) as c FROM attendance_records WHERE session_id=?", (sid,)).fetchone()['c']
        new_d    = db.execute("SELECT COUNT(*) as c FROM attendance_records WHERE session_id=? AND is_new_devotee=1", (sid,)).fetchone()['c']
        return jsonify({'target':target,'present':present,'newDevotees':new_d,'totalPresent':present})


@app.route('/api/sessions/<int:sid>/attendance', methods=['GET'])
def session_attendance(sid):
    with get_db() as db:
        rows = db.execute("""SELECT ar.*,d.name,d.mobile,d.team_name,d.reference_by,d.calling_by,d.chanting_rounds,d.dob,d.devotee_status
                             FROM attendance_records ar JOIN devotees d ON ar.devotee_id=d.id
                             WHERE ar.session_id=? ORDER BY ar.marked_at DESC""", (sid,)).fetchall()
        return jsonify(rows_to_list(rows))


@app.route('/api/sessions/<int:sid>/candidates', methods=['GET'])
def session_candidates(sid):
    search = request.args.get('search', '')
    week   = current_sunday()
    q = """SELECT d.*,cs.coming_status,cs.calling_notes,ar.id as attendance_id
           FROM devotees d
           LEFT JOIN calling_status cs ON d.id=cs.devotee_id AND cs.week_date=?
           LEFT JOIN attendance_records ar ON d.id=ar.devotee_id AND ar.session_id=?
           WHERE d.is_active=1 AND (cs.coming_status IS NULL OR cs.coming_status NOT IN ('Shifted','Not Interested'))"""
    params = [week, sid]
    if search:
        q += " AND (d.name LIKE ? OR d.mobile LIKE ?)"; params += [f'%{search}%', f'%{search}%']
    q += " ORDER BY d.name ASC"
    with get_db() as db:
        return jsonify(rows_to_list(db.execute(q, params).fetchall()))


@app.route('/api/sessions/<int:sid>/attend', methods=['POST'])
def mark_attend(sid):
    data = request.json or {}
    did  = data.get('devotee_id')
    new  = 1 if data.get('is_new_devotee') else 0
    try:
        with get_db() as db:
            db.execute("INSERT INTO attendance_records (session_id,devotee_id,is_new_devotee) VALUES (?,?,?)", (sid, did, new))
            db.execute("UPDATE devotees SET lifetime_attendance=lifetime_attendance+1, inactivity_flag=0 WHERE id=?", (did,))
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Already marked present'}), 409


@app.route('/api/sessions/<int:sid>/attend/<int:did>', methods=['DELETE'])
def undo_attend(sid, did):
    with get_db() as db:
        db.execute("DELETE FROM attendance_records WHERE session_id=? AND devotee_id=?", (sid, did))
        db.execute("UPDATE devotees SET lifetime_attendance=MAX(0,lifetime_attendance-1) WHERE id=?", (did,))
    return jsonify({'success': True})


# ══════════════════════════════════════════════════════════════════════════════
# CALLING STATUS
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/calling', methods=['GET'])
def get_calling():
    week = request.args.get('week') or current_sunday()
    with get_db() as db:
        rows = db.execute("""SELECT d.*,cs.coming_status,cs.calling_notes,cs.called_by,cs.id as calling_id
                             FROM devotees d
                             LEFT JOIN calling_status cs ON d.id=cs.devotee_id AND cs.week_date=?
                             WHERE d.is_active=1 ORDER BY d.name ASC""", (week,)).fetchall()
        return jsonify({'weekDate': week, 'devotees': rows_to_list(rows)})


@app.route('/api/calling/<int:did>', methods=['PUT'])
def update_calling(did):
    data    = request.json or {}
    week    = data.get('week_date') or current_sunday()
    status  = data.get('coming_status', 'Maybe')
    notes   = data.get('calling_notes') or None
    called  = data.get('called_by') or None
    with get_db() as db:
        ex = db.execute("SELECT id FROM calling_status WHERE devotee_id=? AND week_date=?", (did, week)).fetchone()
        if ex:
            db.execute("UPDATE calling_status SET coming_status=?,calling_notes=?,called_by=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",
                       (status, notes, called, ex['id']))
        else:
            db.execute("INSERT INTO calling_status (devotee_id,week_date,coming_status,calling_notes,called_by) VALUES (?,?,?,?,?)",
                       (did, week, status, notes, called))
    return jsonify({'success': True})


# ══════════════════════════════════════════════════════════════════════════════
# REPORTS
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/reports/attendance', methods=['GET'])
def report_attendance():
    session_id = request.args.get('session_id')
    week       = request.args.get('week')
    month      = request.args.get('month')
    with get_db() as db:
        ids = []
        if session_id:
            ids = [int(session_id)]
        elif week:
            end = (date.fromisoformat(week) + timedelta(days=6)).isoformat()
            rows = db.execute("SELECT id FROM attendance_sessions WHERE session_date BETWEEN ? AND ?", (week, end)).fetchall()
            ids = [r['id'] for r in rows]
        elif month:
            rows = db.execute("SELECT id FROM attendance_sessions WHERE strftime('%Y-%m',session_date)=?", (month,)).fetchall()
            ids = [r['id'] for r in rows]
        if not ids:
            s = db.execute("SELECT id FROM attendance_sessions WHERE session_date=?", (today_str(),)).fetchone()
            if s: ids = [s['id']]
        if not ids: return jsonify([])
        ph = ','.join('?' * len(ids))
        rows = db.execute(f"""SELECT d.name,d.mobile,d.chanting_rounds,d.team_name,d.calling_by,
                              ar.is_new_devotee,ar.marked_at,s.session_date
                              FROM attendance_records ar JOIN devotees d ON ar.devotee_id=d.id
                              JOIN attendance_sessions s ON ar.session_id=s.id
                              WHERE ar.session_id IN ({ph}) ORDER BY s.session_date DESC,d.name ASC""", ids).fetchall()
        return jsonify(rows_to_list(rows))


@app.route('/api/reports/serious', methods=['GET'])
def report_serious():
    week = request.args.get('week') or current_sunday()
    sd   = request.args.get('session_date') or today_str()
    teams    = ['Lalita','Vishakha','Tungavidya','Indulekha','Sudevi','Rangadevi','Chitralekha','Champaklata']
    statuses = ['Expected to be Serious','Serious','Most Serious']
    with get_db() as db:
        sess = db.execute("SELECT id FROM attendance_sessions WHERE session_date=?", (sd,)).fetchone()
        sid  = sess['id'] if sess else None
        data = []
        for team in teams:
            for status in statuses:
                promised = db.execute("""SELECT COUNT(*) as c FROM devotees d JOIN calling_status cs ON d.id=cs.devotee_id
                                         WHERE d.team_name=? AND d.devotee_status=? AND cs.week_date=? AND cs.coming_status='Yes'""",
                                      (team, status, week)).fetchone()['c']
                arrived = 0
                if sid:
                    arrived = db.execute("""SELECT COUNT(*) as c FROM attendance_records ar JOIN devotees d ON ar.devotee_id=d.id
                                            WHERE d.team_name=? AND d.devotee_status=? AND ar.session_id=?""",
                                         (team, status, sid)).fetchone()['c']
                data.append({'team':team,'status':status,'promised':promised,'arrived':arrived})
        return jsonify(data)


@app.route('/api/reports/teams', methods=['GET'])
def report_teams():
    week = request.args.get('week') or current_sunday()
    sd   = request.args.get('session_date') or today_str()
    teams = ['Lalita','Vishakha','Tungavidya','Indulekha','Sudevi','Rangadevi','Chitralekha','Champaklata']
    with get_db() as db:
        sess = db.execute("SELECT id FROM attendance_sessions WHERE session_date=?", (sd,)).fetchone()
        sid  = sess['id'] if sess else None
        data = []
        for team in teams:
            total    = db.execute("SELECT COUNT(*) as c FROM devotees WHERE team_name=? AND is_active=1", (team,)).fetchone()['c']
            calling  = db.execute("""SELECT COUNT(*) as c FROM devotees d LEFT JOIN calling_status cs ON d.id=cs.devotee_id AND cs.week_date=?
                                     WHERE d.team_name=? AND d.is_active=1 AND (cs.coming_status IS NULL OR cs.coming_status NOT IN ('Shifted','Not Interested'))""",
                                  (week, team)).fetchone()['c']
            target   = db.execute("""SELECT COUNT(*) as c FROM devotees d JOIN calling_status cs ON d.id=cs.devotee_id
                                     WHERE d.team_name=? AND cs.week_date=? AND cs.coming_status='Yes'""",
                                  (team, week)).fetchone()['c']
            actual   = 0
            if sid:
                actual = db.execute("""SELECT COUNT(*) as c FROM attendance_records ar JOIN devotees d ON ar.devotee_id=d.id
                                       WHERE d.team_name=? AND ar.session_id=?""", (team, sid)).fetchone()['c']
            pct = round(actual / target * 100) if target > 0 else 0
            data.append({'team':team,'total':total,'callingList':calling,'target':target,'actualPresent':actual,'percentage':pct})
        return jsonify(data)


@app.route('/api/reports/trends', methods=['GET'])
def report_trends():
    period = request.args.get('period', 'weekly')
    team   = request.args.get('team', '')
    with get_db() as db:
        if period == 'monthly':
            if team:
                rows = db.execute("""SELECT strftime('%Y-%m',s.session_date) as period, COUNT(DISTINCT ar.devotee_id) as count
                                     FROM attendance_sessions s LEFT JOIN attendance_records ar ON s.id=ar.session_id
                                     LEFT JOIN devotees d ON ar.devotee_id=d.id WHERE d.team_name=?
                                     GROUP BY strftime('%Y-%m',s.session_date) ORDER BY period ASC LIMIT 12""", (team,)).fetchall()
            else:
                rows = db.execute("""SELECT strftime('%Y-%m',s.session_date) as period, COUNT(DISTINCT ar.devotee_id) as count
                                     FROM attendance_sessions s LEFT JOIN attendance_records ar ON s.id=ar.session_id
                                     GROUP BY strftime('%Y-%m',s.session_date) ORDER BY period ASC LIMIT 12""").fetchall()
        else:
            if team:
                rows = db.execute("""SELECT s.session_date as period, COUNT(DISTINCT ar.devotee_id) as count
                                     FROM attendance_sessions s LEFT JOIN attendance_records ar ON s.id=ar.session_id
                                     LEFT JOIN devotees d ON ar.devotee_id=d.id WHERE d.team_name=?
                                     GROUP BY s.session_date ORDER BY period ASC LIMIT 24""", (team,)).fetchall()
            else:
                rows = db.execute("""SELECT s.session_date as period, COUNT(DISTINCT ar.devotee_id) as count
                                     FROM attendance_sessions s LEFT JOIN attendance_records ar ON s.id=ar.session_id
                                     GROUP BY s.session_date ORDER BY period ASC LIMIT 24""").fetchall()
        return jsonify(rows_to_list(rows))


@app.route('/api/reports/export', methods=['GET'])
def export_report():
    sid = request.args.get('session_id')
    if not sid: return jsonify({'error':'session_id required'}), 400
    with get_db() as db:
        sess = db.execute("SELECT * FROM attendance_sessions WHERE id=?", (sid,)).fetchone()
        rows = db.execute("""SELECT d.name as Name, d.mobile as Mobile, d.chanting_rounds as 'Chanting Rounds',
                              d.team_name as Team, d.calling_by as 'Calling By',
                              CASE WHEN ar.is_new_devotee=1 THEN 'New' ELSE 'Regular' END as Type
                              FROM attendance_records ar JOIN devotees d ON ar.devotee_id=d.id
                              WHERE ar.session_id=? ORDER BY d.name ASC""", (sid,)).fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'
    if rows:
        ws.append(list(rows[0].keys()))
        for r in rows:
            ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"attendance_{sess['session_date'] if sess else 'report'}.xlsx"
    return send_file(buf, download_name=fname, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ══════════════════════════════════════════════════════════════════════════════
# DEVOTEE CARE
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/care/absent', methods=['GET'])
def care_absent():
    with get_db() as db:
        sessions = db.execute("SELECT id,session_date FROM attendance_sessions ORDER BY session_date DESC LIMIT 5").fetchall()
        if len(sessions) < 2:
            return jsonify({'absentThisWeek':[],'absentPast2Weeks':[]})
        latest = sessions[0]
        prev4  = sessions[1:5]
        prev2  = sessions[1:3]
        devotees = db.execute("SELECT * FROM devotees WHERE is_active=1").fetchall()
        absent_week, absent_2weeks = [], []
        for d in devotees:
            attended_latest = db.execute("SELECT id FROM attendance_records WHERE session_id=? AND devotee_id=?",
                                         (latest['id'], d['id'])).fetchone()
            if attended_latest: continue
            attended_before = any(db.execute("SELECT id FROM attendance_records WHERE session_id=? AND devotee_id=?",
                                              (s['id'], d['id'])).fetchone() for s in prev4)
            if not attended_before: continue
            missing_prev2 = all(not db.execute("SELECT id FROM attendance_records WHERE session_id=? AND devotee_id=?",
                                                (s['id'], d['id'])).fetchone() for s in prev2)
            (absent_2weeks if missing_prev2 else absent_week).append(row_to_dict(d))
        return jsonify({'absentThisWeek': absent_week, 'absentPast2Weeks': absent_2weeks})


@app.route('/api/care/newcomers', methods=['GET'])
def care_newcomers():
    with get_db() as db:
        sessions = db.execute("SELECT id FROM attendance_sessions ORDER BY session_date DESC LIMIT 2").fetchall()
        if len(sessions) < 2: return jsonify([])
        latest, prev = sessions[0]['id'], sessions[1]['id']
        rows = db.execute("""SELECT d.* FROM devotees d
                             JOIN attendance_records ar1 ON d.id=ar1.devotee_id AND ar1.session_id=? AND ar1.is_new_devotee=1
                             JOIN attendance_records ar2 ON d.id=ar2.devotee_id AND ar2.session_id=?""",
                          (prev, latest)).fetchall()
        return jsonify(rows_to_list(rows))


@app.route('/api/care/birthdays', methods=['GET'])
def care_birthdays():
    today = date.today()
    week_mds = set()
    for i in range(7):
        d = today + timedelta(days=i)
        week_mds.add(f"{d.month:02d}-{d.day:02d}")
    with get_db() as db:
        devotees = db.execute("SELECT * FROM devotees WHERE is_active=1 AND dob IS NOT NULL AND dob!=''").fetchall()
        bdays = [row_to_dict(d) for d in devotees if d['dob'] and d['dob'][5:] in week_mds]
        return jsonify(bdays)


@app.route('/api/care/inactive', methods=['GET'])
def care_inactive():
    update_inactivity_flags()
    with get_db() as db:
        rows = db.execute("SELECT * FROM devotees WHERE inactivity_flag=1 AND is_active=1 ORDER BY name").fetchall()
        return jsonify(rows_to_list(rows))


# ══════════════════════════════════════════════════════════════════════════════
# EVENTS
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/events', methods=['GET'])
def get_events():
    with get_db() as db:
        return jsonify(rows_to_list(db.execute("SELECT * FROM events ORDER BY event_date ASC").fetchall()))


@app.route('/api/events', methods=['POST'])
def create_event():
    data = request.json or {}
    with get_db() as db:
        cur = db.execute("INSERT INTO events (event_name,event_date,description) VALUES (?,?,?)",
                         (data.get('event_name'), data.get('event_date') or None, data.get('description') or None))
        ev = db.execute("SELECT * FROM events WHERE id=?", (cur.lastrowid,)).fetchone()
        return jsonify(row_to_dict(ev)), 201


@app.route('/api/events/<int:eid>', methods=['PUT'])
def update_event(eid):
    data = request.json or {}
    with get_db() as db:
        db.execute("UPDATE events SET event_name=?,event_date=?,description=? WHERE id=?",
                   (data.get('event_name'), data.get('event_date') or None, data.get('description') or None, eid))
    return jsonify({'success': True})


@app.route('/api/events/<int:eid>', methods=['DELETE'])
def delete_event(eid):
    with get_db() as db:
        db.execute("DELETE FROM event_devotees WHERE event_id=?", (eid,))
        db.execute("DELETE FROM events WHERE id=?", (eid,))
    return jsonify({'success': True})


@app.route('/api/events/<int:eid>/devotees', methods=['GET'])
def event_devotees(eid):
    with get_db() as db:
        rows = db.execute("""SELECT d.*,ed.added_at FROM devotees d JOIN event_devotees ed ON d.id=ed.devotee_id
                             WHERE ed.event_id=? ORDER BY d.name ASC""", (eid,)).fetchall()
        return jsonify(rows_to_list(rows))


@app.route('/api/events/<int:eid>/devotees', methods=['POST'])
def add_event_devotee(eid):
    did = (request.json or {}).get('devotee_id')
    try:
        with get_db() as db:
            db.execute("INSERT INTO event_devotees (event_id,devotee_id) VALUES (?,?)", (eid, did))
        return jsonify({'success': True})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Already added'}), 409


@app.route('/api/events/<int:eid>/devotees/<int:did>', methods=['DELETE'])
def remove_event_devotee(eid, did):
    with get_db() as db:
        db.execute("DELETE FROM event_devotees WHERE event_id=? AND devotee_id=?", (eid, did))
    return jsonify({'success': True})


@app.route('/api/events/<int:eid>/export', methods=['GET'])
def export_event(eid):
    with get_db() as db:
        ev   = db.execute("SELECT * FROM events WHERE id=?", (eid,)).fetchone()
        rows = db.execute("""SELECT d.name as Name, d.mobile as Mobile, d.team_name as Team
                             FROM devotees d JOIN event_devotees ed ON d.id=ed.devotee_id
                             WHERE ed.event_id=? ORDER BY d.name ASC""", (eid,)).fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Devotees'
    ws.append(['Name','Mobile','Team'])
    for r in rows: ws.append([r['Name'], r['Mobile'], r['Team']])
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"{(ev['event_name'] if ev else 'event').replace(' ','_')}_devotees.xlsx"
    return send_file(buf, download_name=fname, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── START ─────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("\n  OM - Sakhi Sang Attendance System")
    print("  Running at http://localhost:3000\n")
    app.run(host='0.0.0.0', port=3000, debug=False)
